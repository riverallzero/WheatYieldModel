import datetime
import os.path

import matplotlib
import openpyxl
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt

plt.rcParams['font.family'] = 'NanumGothic'
plt.rcParams['axes.unicode_minus'] = False


def read_LAI(filename: str) -> None:
    doc = openpyxl.load_workbook(filename)

    sheet = doc['LAI']
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    items = ["LAI", "length", "width"]

    datasets = []

    dates = [
        datetime.datetime(2020, 3, 22),
        datetime.datetime(2020, 4, 12),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 14),
        datetime.datetime(2020, 5, 31),
    ]
    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 6 + rep + 3
                        col = i * 9 + j * 3 + 3 + k
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row, column=col).value
                            }
                        )
    return pd.DataFrame(datasets)


def read_SPAD(filename: str) -> None:
    doc = openpyxl.load_workbook(filename)

    sheet = doc["SPAD"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    items = ["SPAD"]

    datasets = []

    dates = [
        datetime.datetime(2020, 3, 22),
        datetime.datetime(2020, 4, 12),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 14),
        datetime.datetime(2020, 5, 31),
    ]
    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 6 + rep + 3
                        col = i * 3 + j * 1 + 3 + k  # 3, 12
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row, column=col).value
                            }
                        )
    return pd.DataFrame(datasets)


def read_Photosynthesis(filename: str) -> None:
    doc = openpyxl.load_workbook(filename)

    sheet = doc["Photosynthesis"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    items = ["F0", "Fm", "Fv", "Tm", "Fv/Fm"]

    datasets = []

    dates = [
        datetime.datetime(2020, 3, 22),
        datetime.datetime(2020, 4, 12),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 14),
        datetime.datetime(2020, 5, 31),
    ]
    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 6 + rep + 3
                        col = i * 15 + j * 5 + k
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 4).value
                            }
                        )
    return pd.DataFrame(datasets)


def read_FreshWeight_0225(filename: str) -> None:
    doc = openpyxl.load_workbook(filename)

    sheet = doc["02.25"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["생체중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 2, 25)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_DryWeight_0225(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["02.25"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["건물중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 2, 25)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 9
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_WaterContent_0225(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["02.25"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["수분함량"]

    datasets = []

    dates = [
        datetime.datetime(2020, 2, 25)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 17
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 1, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_FreshWeight_0322(filename: str) -> None:
    doc = openpyxl.load_workbook(filename)

    sheet = doc["03.22"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["생체중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 3, 22)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_DryWeight_0322(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["03.22"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["건물중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 3, 22)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 9
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_WaterContent_0322(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["03.22"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["수분함량"]

    datasets = []

    dates = [
        datetime.datetime(2020, 3, 22)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 17
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 1, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_FreshWeight_0418(filename: str) -> None:
    doc = openpyxl.load_workbook(filename)

    sheet = doc["04.18"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["생체중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 18)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_DryWeight_0418(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["04.18"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["건물중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 18)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 9
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_WaterContent_0418(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["04.18"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    items = ["수분함량"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 18)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(items):
                    for rep in range(replications):
                        row = d * 0 + rep + 17
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 1, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


# 05.10
def read_seed_FreshWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_DryWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_WaterContent_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_FreshWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_DryWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_WaterContent_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_FreshWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_DryWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_WaterContent_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_FreshWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["생체중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_DryWeight_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["건물중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 8 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_WaterContent_0510(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.10"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["수분함량"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 10)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 15 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


# 05.31
def read_seed_FreshWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_DryWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_WaterContent_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_FreshWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_DryWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_WaterContent_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_FreshWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_DryWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_WaterContent_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_FreshWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["생체중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_DryWeight_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["건물중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 8 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_WaterContent_0531(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["05.31"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["수분함량"]

    datasets = []

    dates = [
        datetime.datetime(2020, 5, 31)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 15 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


# 06.17
def read_seed_FreshWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_DryWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_WaterContent_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_FreshWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_DryWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_WaterContent_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_FreshWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_DryWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_WaterContent_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_FreshWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["생체중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_DryWeight_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["건물중"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 8 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_200_WaterContent_0617(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["06.17"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 3
    item1 = ["수분함량"]

    datasets = []

    dates = [
        datetime.datetime(2020, 6, 17)
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 15 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


# 개화기
def read_seed_FreshWeight_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_DryWeight_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_seed_WaterContent_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_FreshWeight_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_DryWeight_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_leaf_WaterContent_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 29
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_FreshWeight_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["생체중(면적당)"]
    item2 = ["생체중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 3, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_DryWeight_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    item2 = ["건물중(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 10 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_stem_WaterContent_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["수분함량(면적당)"]
    item2 = ["수분함량(개체당)"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 1).value
                            }
                        )

                for k, item in enumerate(item2):
                    for rep in range(replications):
                        row = d * 0 + rep + 19 + 59
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 2, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_root_DryWeight_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 5
    item1 = ["건물중(면적당)"]
    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 4, column=col + 16).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_1_bloom(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["개화기"]
    groups = ["A(한발)", "B(무처리)", "C(적정)", "D(과습)"]
    blocks = 3
    replications = 10
    item1 = ["1수영화수"]

    datasets = []

    dates = [
        datetime.datetime(2020, 4, 28),
        datetime.datetime(2020, 4, 29),
        datetime.datetime(2020, 4, 30),
        datetime.datetime(2020, 5, 1),
        datetime.datetime(2020, 5, 2),
        datetime.datetime(2020, 5, 3),
    ]

    for d, date in enumerate(dates):
        for i, group in enumerate(groups):
            for j in range(blocks):
                block = j + 1
                for k, item in enumerate(item1):
                    for rep in range(replications):
                        row = d * 0 + rep + 1 + 88
                        col = i * 3 + j * 1 + k + 2
                        datasets.append(
                            {
                                "date": date, "group": group, "block": block, "rep": rep + 1, "item": item,
                                "value": sheet.cell(row=row + 4, column=col + 1).value
                            }
                        )

    return pd.DataFrame(datasets)


def read_yield(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)

    sheet = doc["Yield"]
    group1 = ["A(한발)"]
    group2 = ["B(무처리)"]
    group3 = ["C(적정)"]
    group4 = ["D(과습)"]

    replications = 3
    item1 = ["G(kg)"]
    item2 = ["Yield"]

    # Sa(m^2)
    Sa = 0.7225
    # S'a(m^2)
    SSm = 660
    # S''a(ha)
    SSha = 0.229

    datasets = []

    for i, group in enumerate(group1):
        for k, item in enumerate(item1):
            for rep in range(replications):
                row = rep + 4
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )

            for rep in range(replications):
                row = rep + 7
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
            for rep in range(replications):
                row = rep + 10
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
        for k, item in enumerate(item2):
            for rep in range(replications):
                row = rep + 4
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

            for rep in range(replications):
                row = rep + 7
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

            for rep in range(replications):
                row = rep + 10
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

    for i, group in enumerate(group2):
        for k, item in enumerate(item1):
            for rep in range(replications):
                row = rep + 16
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
            for rep in range(replications):
                row = rep + 19
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
            for rep in range(replications):
                row = rep + 22
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
        for k, item in enumerate(item2):
            for rep in range(replications):
                row = rep + 16
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )
            for rep in range(replications):
                row = rep + 19
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

            for rep in range(replications):
                row = rep + 22
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

    for i, group in enumerate(group3):
        for k, item in enumerate(item1):
            for rep in range(replications):
                row = rep + 28
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
            for rep in range(replications):
                row = rep + 31
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
            for rep in range(replications):
                row = rep + 34
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
        for k, item in enumerate(item2):
            for rep in range(replications):
                row = rep + 28
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )
            for rep in range(replications):
                row = rep + 31
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

            for rep in range(replications):
                row = rep + 34
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

    for i, group in enumerate(group4):
        for k, item in enumerate(item1):
            for rep in range(replications):
                row = rep + 40
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
            for rep in range(replications):
                row = rep + 43
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
            for rep in range(replications):
                row = rep + 46
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": sheet.cell(row=row, column=col).value
                    }
                )
        for k, item in enumerate(item2):
            for rep in range(replications):
                row = rep + 40
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 1, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )
            for rep in range(replications):
                row = rep + 43
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 2, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

            for rep in range(replications):
                row = rep + 46
                col = i * 9 + k + 2
                datasets.append(
                    {
                        "group": group, "block": 3, "rep": rep + 1, "item": item,
                        "value": (sheet.cell(row=row, column=col).value) / Sa * SSm / SSha
                    }
                )

    return pd.DataFrame(datasets)


def save_unbong():
    output_dir = "../output/unbong/"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    filename = "../input/20220711_Unbong데이터_updated.xlsx"

    # LAI
    df_LAI = read_LAI(filename)
    df_LAI.to_csv(os.path.join(output_dir, "U_LAI.csv"), index=False, encoding="utf-8-sig")
    df_LAI['date_str'] = df_LAI['date'].dt.strftime('%Y/%m/%d')

    # SPAD
    df_SPAD = read_SPAD(filename)
    df_SPAD.to_csv(os.path.join(output_dir, "U_SPAD.csv"), index=False, encoding="utf-8-sig")
    df_SPAD['date_str'] = df_SPAD['date'].dt.strftime('%Y/%m/%d')

    # Photosynthesis
    df_Photosynthesis = read_Photosynthesis(filename)
    # df_Photosynthesis = df_Photosynthesis.dropna()
    df_Photosynthesis.to_csv(os.path.join(output_dir, "U_Photosynthesis.csv"), index=False, encoding="utf-8-sig")
    df_Photosynthesis['date_str'] = df_Photosynthesis['date'].dt.strftime('%Y/%m/%d')

    # FreshWeight_0225
    df_FreshWeight_0225 = read_FreshWeight_0225(filename)
    df_FreshWeight_0225.to_csv(os.path.join(output_dir, "U_FreshWeight_0225.csv"), index=False, encoding="utf-8-sig")
    df_FreshWeight_0225['date_str'] = df_FreshWeight_0225['date'].dt.strftime('%Y/%m/%d')

    # DryWeight_0225
    df_DryWeight_0225 = read_DryWeight_0225(filename)
    df_DryWeight_0225.to_csv(os.path.join(output_dir, "U_DryWeight_0225.csv"), index=False, encoding="utf-8-sig")
    df_DryWeight_0225['date_str'] = df_DryWeight_0225['date'].dt.strftime('%Y/%m/%d')

    # WaterContent_0225
    df_WaterContent_0225 = read_WaterContent_0225(filename)
    df_WaterContent_0225.to_csv(os.path.join(output_dir, "U_WaterContent_0225.csv"), index=False, encoding="utf-8-sig")
    df_WaterContent_0225['date_str'] = df_WaterContent_0225['date'].dt.strftime('%Y/%m/%d')

    # FreshWeight_0322
    df_FreshWeight_0322 = read_FreshWeight_0322(filename)
    df_FreshWeight_0322.to_csv(os.path.join(output_dir, "U_FreshWeight_0322.csv"), index=False, encoding="utf-8-sig")
    df_FreshWeight_0322['date_str'] = df_FreshWeight_0322['date'].dt.strftime('%Y/%m/%d')

    # DryWeight_0322
    df_DryWeight_0322 = read_DryWeight_0322(filename)
    df_DryWeight_0322.to_csv(os.path.join(output_dir, "U_DryWeight_0322.csv"), index=False, encoding="utf-8-sig")
    df_DryWeight_0322['date_str'] = df_DryWeight_0322['date'].dt.strftime('%Y/%m/%d')

    # WaterContent_0322
    df_WaterContent_0322 = read_WaterContent_0322(filename)
    df_WaterContent_0322.to_csv(os.path.join(output_dir, "U_WaterContent_0322.csv"), index=False, encoding="utf-8-sig")
    df_WaterContent_0322['date_str'] = df_WaterContent_0322['date'].dt.strftime('%Y/%m/%d')

    # FreshWeight_0418
    df_FreshWeight_0418 = read_FreshWeight_0418(filename)
    # df_FreshWeight_0418 = df_FreshWeight_0418[df_FreshWeight_0418 != "-"]
    # df_FreshWeight_0418 = df_FreshWeight_0418.dropna()
    df_FreshWeight_0418.to_csv(os.path.join(output_dir, "U_FreshWeight_0418.csv"), index=False, encoding="utf-8-sig")
    df_FreshWeight_0418['date_str'] = df_FreshWeight_0418['date'].dt.strftime('%Y/%m/%d')

    # DryWeight_0418
    df_DryWeight_0418 = read_DryWeight_0418(filename)
    # df_DryWeight_0418 = df_DryWeight_0418[df_DryWeight_0418 != "-"]
    # df_DryWeight_0418 = df_DryWeight_0418.dropna()
    df_DryWeight_0418.to_csv(os.path.join(output_dir, "U_DryWeight_0418.csv"), index=False, encoding="utf-8-sig")
    df_DryWeight_0418['date_str'] = df_DryWeight_0418['date'].dt.strftime('%Y/%m/%d')

    # WaterContent_0418
    df_WaterContent_0418 = read_WaterContent_0418(filename)
    # df_WaterContent_0418 = df_WaterContent_0418.dropna()
    df_WaterContent_0418.to_csv(os.path.join(output_dir, "U_WaterContent_0418.csv"), index=False, encoding="utf-8-sig")
    df_WaterContent_0418['date_str'] = df_WaterContent_0418['date'].dt.strftime('%Y/%m/%d')

    # seed_FreshWeight_0510
    df_seed_FreshWeight_0510 = read_seed_FreshWeight_0510(filename)
    # df_seed_FreshWeight_0510 = df_seed_FreshWeight_0510.dropna()
    df_seed_FreshWeight_0510.to_csv(os.path.join(output_dir, "U_seed_FreshWeight_0510.csv"), index=False,
                                    encoding="utf-8-sig")
    df_seed_FreshWeight_0510['date_str'] = df_seed_FreshWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # seed_DryWeight_0510
    df_seed_DryWeight_0510 = read_seed_DryWeight_0510(filename)
    # df_seed_DryWeight_0510 = df_seed_DryWeight_0510.dropna()
    df_seed_DryWeight_0510.to_csv(os.path.join(output_dir, "U_seed_DryWeight_0510.csv"), index=False,
                                  encoding="utf-8-sig")
    df_seed_DryWeight_0510['date_str'] = df_seed_DryWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # seed_WaterContent_0510
    df_seed_WaterContent_0510 = read_seed_WaterContent_0510(filename)
    # df_seed_WaterContent_0510 = df_seed_WaterContent_0510.dropna()
    df_seed_WaterContent_0510.to_csv(os.path.join(output_dir, "U_seed_WaterContent_0510.csv"), index=False,
                                     encoding="utf-8-sig")
    df_seed_WaterContent_0510['date_str'] = df_seed_WaterContent_0510['date'].dt.strftime('%Y/%m/%d')

    # leaf_FreshWeight_0510
    df_leaf_FreshWeight_0510 = read_leaf_FreshWeight_0510(filename)
    # df_leaf_FreshWeight_0510 = df_leaf_FreshWeight_0510.dropna()
    df_leaf_FreshWeight_0510.to_csv(os.path.join(output_dir, "U_leaf_FreshWeight_0510.csv"), index=False,
                                    encoding="utf-8-sig")
    df_leaf_FreshWeight_0510['date_str'] = df_leaf_FreshWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # leaf_DryWeight_0510
    df_leaf_DryWeight_0510 = read_leaf_DryWeight_0510(filename)
    # df_leaf_DryWeight_0510 = df_leaf_DryWeight_0510.dropna()
    df_leaf_DryWeight_0510.to_csv(os.path.join(output_dir, "U_leaf_DryWeight_0510.csv"), index=False,
                                  encoding="utf-8-sig")
    df_leaf_DryWeight_0510['date_str'] = df_leaf_DryWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # leaf_WaterContent_0510
    df_leaf_WaterContent_0510 = read_leaf_WaterContent_0510(filename)
    # df_leaf_WaterContent_0510 = df_leaf_WaterContent_0510.dropna()
    df_leaf_WaterContent_0510.to_csv(os.path.join(output_dir, "U_leaf_WaterContent_0510.csv"), index=False,
                                     encoding="utf-8-sig")
    df_leaf_WaterContent_0510['date_str'] = df_leaf_WaterContent_0510['date'].dt.strftime('%Y/%m/%d')

    # stem_FreshWeight_0510
    df_stem_FreshWeight_0510 = read_stem_FreshWeight_0510(filename)
    # df_stem_FreshWeight_0510 = df_stem_FreshWeight_0510.dropna()
    df_stem_FreshWeight_0510.to_csv(os.path.join(output_dir, "U_stem_FreshWeight_0510.csv"), index=False,
                                    encoding="utf-8-sig")
    df_stem_FreshWeight_0510['date_str'] = df_stem_FreshWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # stem_DryWeight_0510
    df_stem_DryWeight_0510 = read_stem_DryWeight_0510(filename)
    # df_stem_DryWeight_0510 = df_stem_DryWeight_0510.dropna()
    df_stem_DryWeight_0510.to_csv(os.path.join(output_dir, "U_stem_DryWeight_0510.csv"), index=False,
                                  encoding="utf-8-sig")
    df_stem_DryWeight_0510['date_str'] = df_stem_DryWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # stem_WaterContent_0510
    df_stem_WaterContent_0510 = read_stem_WaterContent_0510(filename)
    # df_stem_WaterContent_0510 = df_stem_WaterContent_0510.dropna()
    df_stem_WaterContent_0510.to_csv(os.path.join(output_dir, "U_stem_WaterContent_0510.csv"), index=False,
                                     encoding="utf-8-sig")
    df_stem_WaterContent_0510['date_str'] = df_stem_WaterContent_0510['date'].dt.strftime('%Y/%m/%d')

    # 200_FreshWeight_0510
    df_200_FreshWeight_0510 = read_200_FreshWeight_0510(filename)
    # df_200_FreshWeight_0510 = df_200_FreshWeight_0510.dropna()
    df_200_FreshWeight_0510.to_csv(os.path.join(output_dir, "U_200_FreshWeight_0510.csv"), index=False,
                                   encoding="utf-8-sig")
    df_200_FreshWeight_0510['date_str'] = df_200_FreshWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # 200_DryWeight_0510
    df_200_DryWeight_0510 = read_200_DryWeight_0510(filename)
    # df_200_DryWeight_0510 = df_200_DryWeight_0510.dropna()
    df_200_DryWeight_0510.to_csv(os.path.join(output_dir, "U_200_DryWeight_0510.csv"), index=False,
                                 encoding="utf-8-sig")
    df_200_DryWeight_0510['date_str'] = df_200_DryWeight_0510['date'].dt.strftime('%Y/%m/%d')

    # 200_WaterContent_0510
    df_200_WaterContent_0510 = read_200_WaterContent_0510(filename)
    # df_200_WaterContent_0510 = df_200_WaterContent_0510.dropna()
    df_200_WaterContent_0510.to_csv(os.path.join(output_dir, "U_200_WaterContent_0510.csv"), index=False,
                                    encoding="utf-8-sig")
    df_200_WaterContent_0510['date_str'] = df_200_WaterContent_0510['date'].dt.strftime('%Y/%m/%d')

    # seed_FreshWeight_0531
    df_seed_FreshWeight_0531 = read_seed_FreshWeight_0531(filename)
    # df_seed_FreshWeight_0531 = df_seed_FreshWeight_0531.dropna()
    df_seed_FreshWeight_0531.to_csv(os.path.join(output_dir, "U_seed_FreshWeight_0531.csv"), index=False,
                                    encoding="utf-8-sig")
    df_seed_FreshWeight_0531['date_str'] = df_seed_FreshWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # seed_DryWeight_0531
    df_seed_DryWeight_0531 = read_seed_DryWeight_0531(filename)
    # df_seed_DryWeight_0531 = df_seed_DryWeight_0531.dropna()
    df_seed_DryWeight_0531.to_csv(os.path.join(output_dir, "U_seed_DryWeight_0531.csv"), index=False,
                                  encoding="utf-8-sig")
    df_seed_DryWeight_0531['date_str'] = df_seed_DryWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # seed_WaterContent_0531
    df_seed_WaterContent_0531 = read_seed_WaterContent_0531(filename)
    # df_seed_WaterContent_0531 = df_seed_WaterContent_0531.dropna()
    df_seed_WaterContent_0531.to_csv(os.path.join(output_dir, "U_seed_WaterContent_0531.csv"), index=False,
                                     encoding="utf-8-sig")
    df_seed_WaterContent_0531['date_str'] = df_seed_WaterContent_0531['date'].dt.strftime('%Y/%m/%d')

    # leaf_FreshWeight_0531
    df_leaf_FreshWeight_0531 = read_leaf_FreshWeight_0531(filename)
    # df_leaf_FreshWeight_0531 = df_leaf_FreshWeight_0531.dropna()
    df_leaf_FreshWeight_0531.to_csv(os.path.join(output_dir, "U_leaf_FreshWeight_0531.csv"), index=False,
                                    encoding="utf-8-sig")
    df_leaf_FreshWeight_0531['date_str'] = df_leaf_FreshWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # leaf_DryWeight_0531
    df_leaf_DryWeight_0531 = read_leaf_DryWeight_0531(filename)
    # df_leaf_DryWeight_0531 = df_leaf_DryWeight_0531.dropna()
    df_leaf_DryWeight_0531.to_csv(os.path.join(output_dir, "U_leaf_DryWeight_0531.csv"), index=False,
                                  encoding="utf-8-sig")
    df_leaf_DryWeight_0531['date_str'] = df_leaf_DryWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # leaf_WaterContent_0531
    df_leaf_WaterContent_0531 = read_leaf_WaterContent_0531(filename)
    # df_leaf_WaterContent_0531 = df_leaf_WaterContent_0531.dropna()
    df_leaf_WaterContent_0531.to_csv(os.path.join(output_dir, "U_leaf_WaterContent_0531.csv"), index=False,
                                     encoding="utf-8-sig")
    df_leaf_WaterContent_0531['date_str'] = df_leaf_WaterContent_0531['date'].dt.strftime('%Y/%m/%d')

    # stem_FreshWeight_0531
    df_stem_FreshWeight_0531 = read_stem_FreshWeight_0531(filename)
    # df_stem_FreshWeight_0531 = df_stem_FreshWeight_0531.dropna()
    df_stem_FreshWeight_0531.to_csv(os.path.join(output_dir, "U_stem_FreshWeight_0531.csv"), index=False,
                                    encoding="utf-8-sig")
    df_stem_FreshWeight_0531['date_str'] = df_stem_FreshWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # stem_DryWeight_0531
    df_stem_DryWeight_0531 = read_stem_DryWeight_0531(filename)
    # df_stem_DryWeight_0531 = df_stem_DryWeight_0531.dropna()
    df_stem_DryWeight_0531.to_csv(os.path.join(output_dir, "U_stem_DryWeight_0531.csv"), index=False,
                                  encoding="utf-8-sig")
    df_stem_DryWeight_0531['date_str'] = df_stem_DryWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # stem_WaterContent_0531
    df_stem_WaterContent_0531 = read_stem_WaterContent_0531(filename)
    # df_stem_WaterContent_0531 = df_stem_WaterContent_0531.dropna()
    df_stem_WaterContent_0531.to_csv(os.path.join(output_dir, "U_stem_WaterContent_0531.csv"), index=False,
                                     encoding="utf-8-sig")
    df_stem_WaterContent_0531['date_str'] = df_stem_WaterContent_0531['date'].dt.strftime('%Y/%m/%d')

    # 200_FreshWeight_0531
    df_200_FreshWeight_0531 = read_200_FreshWeight_0531(filename)
    # df_200_FreshWeight_0531 = df_200_FreshWeight_0531.dropna()
    df_200_FreshWeight_0531.to_csv(os.path.join(output_dir, "U_200_FreshWeight_0531.csv"), index=False,
                                   encoding="utf-8-sig")
    df_200_FreshWeight_0531['date_str'] = df_200_FreshWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # 200_DryWeight_0531
    df_200_DryWeight_0531 = read_200_DryWeight_0531(filename)
    # df_200_DryWeight_0531 = df_200_DryWeight_0531.dropna()
    df_200_DryWeight_0531.to_csv(os.path.join(output_dir, "U_200_DryWeight_0531.csv"), index=False,
                                 encoding="utf-8-sig")
    df_200_DryWeight_0531['date_str'] = df_200_DryWeight_0531['date'].dt.strftime('%Y/%m/%d')

    # 200_WaterContent_0531
    df_200_WaterContent_0531 = read_200_WaterContent_0531(filename)
    # df_200_WaterContent_0531 = df_200_WaterContent_0531.dropna()
    df_200_WaterContent_0531.to_csv(os.path.join(output_dir, "U_200_WaterContent_0531.csv"), index=False,
                                    encoding="utf-8-sig")
    df_200_WaterContent_0531['date_str'] = df_200_WaterContent_0531['date'].dt.strftime('%Y/%m/%d')

    # seed_FreshWeight_0617
    df_seed_FreshWeight_0617 = read_seed_FreshWeight_0617(filename)
    # df_seed_FreshWeight_0617 = df_seed_FreshWeight_0617.dropna()
    df_seed_FreshWeight_0617.to_csv(os.path.join(output_dir, "U_seed_FreshWeight_0617.csv"), index=False,
                                    encoding="utf-8-sig")
    df_seed_FreshWeight_0617['date_str'] = df_seed_FreshWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # seed_DryWeight_0617
    df_seed_DryWeight_0617 = read_seed_DryWeight_0617(filename)
    # df_seed_DryWeight_0617 = df_seed_DryWeight_0617.dropna()
    df_seed_DryWeight_0617.to_csv(os.path.join(output_dir, "U_seed_DryWeight_0617.csv"), index=False,
                                  encoding="utf-8-sig")
    df_seed_DryWeight_0617['date_str'] = df_seed_DryWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # seed_WaterContent_0617
    df_seed_WaterContent_0617 = read_seed_WaterContent_0617(filename)
    # df_seed_WaterContent_0617 = df_seed_WaterContent_0617.dropna()
    df_seed_WaterContent_0617.to_csv(os.path.join(output_dir, "U_seed_WaterContent_0617.csv"), index=False,
                                     encoding="utf-8-sig")
    df_seed_WaterContent_0617['date_str'] = df_seed_WaterContent_0617['date'].dt.strftime('%Y/%m/%d')

    # leaf_FreshWeight_0617
    df_leaf_FreshWeight_0617 = read_leaf_FreshWeight_0617(filename)
    # df_leaf_FreshWeight_0617 = df_leaf_FreshWeight_0617.dropna()
    df_leaf_FreshWeight_0617.to_csv(os.path.join(output_dir, "U_leaf_FreshWeight_0617.csv"), index=False,
                                    encoding="utf-8-sig")
    df_leaf_FreshWeight_0617['date_str'] = df_leaf_FreshWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # leaf_DryWeight_0617
    df_leaf_DryWeight_0617 = read_leaf_DryWeight_0617(filename)
    # df_leaf_DryWeight_0617 = df_leaf_DryWeight_0617.dropna()
    df_leaf_DryWeight_0617.to_csv(os.path.join(output_dir, "U_leaf_DryWeight_0617.csv"), index=False,
                                  encoding="utf-8-sig")
    df_leaf_DryWeight_0617['date_str'] = df_leaf_DryWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # leaf_WaterContent_0617
    df_leaf_WaterContent_0617 = read_leaf_WaterContent_0617(filename)
    # df_leaf_WaterContent_0617 = df_leaf_WaterContent_0617.dropna()
    df_leaf_WaterContent_0617.to_csv(os.path.join(output_dir, "U_leaf_WaterContent_0617.csv"), index=False,
                                     encoding="utf-8-sig")
    df_leaf_WaterContent_0617['date_str'] = df_leaf_WaterContent_0617['date'].dt.strftime('%Y/%m/%d')

    # stem_FreshWeight_0617
    df_stem_FreshWeight_0617 = read_stem_FreshWeight_0617(filename)
    # df_stem_FreshWeight_0617 = df_stem_FreshWeight_0617.dropna()
    df_stem_FreshWeight_0617.to_csv(os.path.join(output_dir, "U_stem_FreshWeight_0617.csv"), index=False,
                                    encoding="utf-8-sig")
    df_stem_FreshWeight_0617['date_str'] = df_stem_FreshWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # stem_DryWeight_0617
    df_stem_DryWeight_0617 = read_stem_DryWeight_0617(filename)
    # df_stem_DryWeight_0617 = df_stem_DryWeight_0617.dropna()
    df_stem_DryWeight_0617.to_csv(os.path.join(output_dir, "U_stem_DryWeight_0617.csv"), index=False,
                                  encoding="utf-8-sig")
    df_stem_DryWeight_0617['date_str'] = df_stem_DryWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # stem_WaterContent_0617
    df_stem_WaterContent_0617 = read_stem_WaterContent_0617(filename)
    # df_stem_WaterContent_0617 = df_stem_WaterContent_0617.dropna()
    df_stem_WaterContent_0617.to_csv(os.path.join(output_dir, "U_stem_WaterContent_0617.csv"), index=False,
                                     encoding="utf-8-sig")
    df_stem_WaterContent_0617['date_str'] = df_stem_WaterContent_0617['date'].dt.strftime('%Y/%m/%d')

    # 200_FreshWeight_0617
    df_200_FreshWeight_0617 = read_200_FreshWeight_0617(filename)
    # df_200_FreshWeight_0617 = df_200_FreshWeight_0617.dropna()
    df_200_FreshWeight_0617.to_csv(os.path.join(output_dir, "U_200_FreshWeight_0617.csv"), index=False,
                                   encoding="utf-8-sig")
    df_200_FreshWeight_0617['date_str'] = df_200_FreshWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # 200_DryWeight_0617
    df_200_DryWeight_0617 = read_200_DryWeight_0617(filename)
    # df_200_DryWeight_0617 = df_200_DryWeight_0617.dropna()
    df_200_DryWeight_0617.to_csv(os.path.join(output_dir, "U_200_DryWeight_0617.csv"), index=False,
                                 encoding="utf-8-sig")
    df_200_DryWeight_0617['date_str'] = df_200_DryWeight_0617['date'].dt.strftime('%Y/%m/%d')

    # 200_WaterContent_0617
    df_200_WaterContent_0617 = read_200_WaterContent_0617(filename)
    # df_200_WaterContent_0617 = df_200_WaterContent_0617.dropna()
    df_200_WaterContent_0617.to_csv(os.path.join(output_dir, "U_200_WaterContent_0617.csv"), index=False,
                                    encoding="utf-8-sig")
    df_200_WaterContent_0617['date_str'] = df_200_WaterContent_0617['date'].dt.strftime('%Y/%m/%d')

    # seed_FreshWeight_bloom
    df_seed_FreshWeight_bloom = read_seed_FreshWeight_bloom(filename)
    # df_seed_FreshWeight_bloom = df_seed_FreshWeight_bloom.dropna()
    df_seed_FreshWeight_bloom.to_csv(os.path.join(output_dir, "U_seed_FreshWeight_bloom.csv"), index=False,
                                     encoding="utf-8-sig")
    df_seed_FreshWeight_bloom['date_str'] = df_seed_FreshWeight_bloom['date'].dt.strftime('%Y/%m/%d')

    # seed_DryWeight_bloom
    df_seed_DryWeight_bloom = read_seed_DryWeight_bloom(filename)
    # df_seed_DryWeight_bloom = df_seed_DryWeight_bloom.dropna()
    df_seed_DryWeight_bloom.to_csv(os.path.join(output_dir, "U_seed_DryWeight_bloom.csv"), index=False,
                                   encoding="utf-8-sig")
    df_seed_DryWeight_bloom['date_str'] = df_seed_DryWeight_bloom['date'].dt.strftime('%Y/%m/%d')

    # seed_WaterContent_bloom
    df_seed_WaterContent_bloom = read_seed_WaterContent_bloom(filename)
    # df_seed_WaterContent_bloom = df_seed_WaterContent_bloom.dropna()
    df_seed_WaterContent_bloom.to_csv(os.path.join(output_dir, "U_seed_WaterContent_bloom.csv"), index=False,
                                      encoding="utf-8-sig")
    df_seed_WaterContent_bloom['date_str'] = df_seed_WaterContent_bloom['date'].dt.strftime('%Y/%m/%d')

    # leaf_FreshWeight_bloom
    df_leaf_FreshWeight_bloom = read_leaf_FreshWeight_bloom(filename)
    # df_leaf_FreshWeight_bloom = df_leaf_FreshWeight_bloom[df_leaf_FreshWeight_bloom != "-"]
    # df_leaf_FreshWeight_bloom = df_leaf_FreshWeight_bloom.dropna()
    df_leaf_FreshWeight_bloom.to_csv(os.path.join(output_dir, "U_leaf_FreshWeight_bloom.csv"), index=False,
                                     encoding="utf-8-sig")
    df_leaf_FreshWeight_bloom['date_str'] = df_leaf_FreshWeight_bloom['date'].dt.strftime('%Y/%m/%d')

    # leaf_DryWeight_bloom
    df_leaf_DryWeight_bloom = read_leaf_DryWeight_bloom(filename)
    # df_leaf_DryWeight_bloom = df_leaf_DryWeight_bloom[df_leaf_DryWeight_bloom != "-"]
    # df_leaf_DryWeight_bloom = df_leaf_DryWeight_bloom.dropna()
    df_leaf_DryWeight_bloom.to_csv(os.path.join(output_dir, "U_leaf_DryWeight_bloom.csv"), index=False,
                                   encoding="utf-8-sig")
    df_leaf_DryWeight_bloom['date_str'] = df_leaf_DryWeight_bloom['date'].dt.strftime('%Y/%m/%d')

    # leaf_WaterContent_bloom
    df_leaf_WaterContent_bloom = read_leaf_WaterContent_bloom(filename)
    # df_leaf_WaterContent_bloom = df_leaf_WaterContent_bloom[df_leaf_WaterContent_bloom != "-"]
    # df_leaf_WaterContent_bloom = df_leaf_WaterContent_bloom.dropna()
    df_leaf_WaterContent_bloom.to_csv(os.path.join(output_dir, "U_leaf_WaterContent_bloom.csv"), index=False,
                                      encoding="utf-8-sig")
    df_leaf_WaterContent_bloom['date_str'] = df_leaf_WaterContent_bloom['date'].dt.strftime('%Y/%m/%d')

    # stem_FreshWeight_bloom
    df_stem_FreshWeight_bloom = read_stem_FreshWeight_bloom(filename)
    # df_stem_FreshWeight_bloom = df_stem_FreshWeight_bloom[df_stem_FreshWeight_bloom != "X"]
    # df_stem_FreshWeight_bloom = df_stem_FreshWeight_bloom[df_stem_FreshWeight_bloom != "-"]
    #
    # df_stem_FreshWeight_bloom = df_stem_FreshWeight_bloom.dropna()
    df_stem_FreshWeight_bloom.to_csv(os.path.join(output_dir, "U_stem_FreshWeight_bloom.csv"), index=False,
                                     encoding="utf-8-sig")
    df_stem_FreshWeight_bloom['date_str'] = df_stem_FreshWeight_bloom['date'].dt.strftime('%Y/%m/%d')

    # stem_DryWeight_bloom
    df_stem_DryWeight_bloom = read_stem_DryWeight_bloom(filename)
    # df_stem_DryWeight_bloom = df_stem_DryWeight_bloom[df_stem_DryWeight_bloom != "X"]
    # df_stem_DryWeight_bloom = df_stem_DryWeight_bloom[df_stem_DryWeight_bloom != "-"]
    # df_stem_DryWeight_bloom = df_stem_DryWeight_bloom.dropna()
    df_stem_DryWeight_bloom.to_csv(os.path.join(output_dir, "U_stem_DryWeight_bloom.csv"), index=False,
                                   encoding="utf-8-sig")
    df_stem_DryWeight_bloom['date_str'] = df_stem_DryWeight_bloom['date'].dt.strftime('%Y/%m/%d')

    # stem_WaterContent_bloom
    df_stem_WaterContent_bloom = read_stem_WaterContent_bloom(filename)
    # df_stem_WaterContent_bloom = df_stem_WaterContent_bloom.dropna()
    df_stem_WaterContent_bloom.to_csv(os.path.join(output_dir, "U_stem_WaterContent_bloom.csv"), index=False,
                                      encoding="utf-8-sig")
    df_stem_WaterContent_bloom['date_str'] = df_stem_WaterContent_bloom['date'].dt.strftime('%Y/%m/%d')

    # root_DryWeight_bloom
    df_root_DryWeight_bloom = read_root_DryWeight_bloom(filename)
    # df_root_DryWeight_bloom = df_root_DryWeight_bloom[df_root_DryWeight_bloom != "X"]
    # df_root_DryWeight_bloom = df_root_DryWeight_bloom[df_root_DryWeight_bloom != "-"]
    # df_root_DryWeight_bloom = df_root_DryWeight_bloom.dropna()
    df_root_DryWeight_bloom.to_csv(os.path.join(output_dir, "U_root_DryWeight_bloom.csv"), index=False,
                                   encoding="utf-8-sig")
    df_root_DryWeight_bloom['date_str'] = df_root_DryWeight_bloom['date'].dt.strftime('%Y/%m/%d')

    # 1_bloom
    df_1_bloom = read_1_bloom(filename)
    # df_1_bloom = df_1_bloom[df_1_bloom != "X"]
    # df_1_bloom = df_1_bloom[df_1_bloom != "-"]
    # df_1_bloom = df_1_bloom.dropna()
    df_1_bloom.to_csv(os.path.join(output_dir, "U_1_bloom.csv"), index=False,
                      encoding="utf-8-sig")
    df_1_bloom['date_str'] = df_1_bloom['date'].dt.strftime('%Y/%m/%d')

    # Yield
    df_yield = read_yield(filename)
    df_yield.to_csv(os.path.join(output_dir, "U_Yield.csv"), index=False,
                    encoding="utf-8-sig")


def main():
    save_unbong()


if __name__ == "__main__":
    main()